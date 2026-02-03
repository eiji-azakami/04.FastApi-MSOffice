"""
Excel出力API

このモジュールは、Excelファイル出力のサンプルプログラムです。
"""

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from io import BytesIO

router = APIRouter(tags=["excel"])


@router.get("/excel/export")
def export_excel():
    """
    Excelファイル出力
    """
    # book生成
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "所得の内訳書"

    # タイトル部分
    title_cell = sheet["C1"]
    title_cell.value = "所得の内訳書"
    sheet.row_dimensions[1].height = 35.0
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True, size=18)
    sheet.merge_cells("C1:G1")

    # 住所と氏名部分
    cell = sheet["E2"]
    cell.value = "住所"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell = sheet["E3"]
    cell.value = "氏名"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    for row in [2, 3]:
        cell = sheet[f"F{row}"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))
        cell = sheet[f"G{row}"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))

    sheet.row_dimensions[2].height = 17.0
    sheet.row_dimensions[3].height = 17.0

    sheet.row_dimensions[4].height = 9.0

    # 表のヘッダー部分
    headers = [
        "所得の\n種類",
        "種目",
        "所得の生ずる場所又は給与などの支払者の\n氏名・名称・住所・所在地・法人番号・電話番号",
        "所得の基因となる資産の数量",
        "収入金額（源泉徴収税額を含む）",
        "源泉徴収税額",
        "支払を受けた年月",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # 表の内容部分
    for row_idx in range(6, 16):
        sheet.row_dimensions[row_idx].height = 49  # 行高さを49に設定
        for col_idx in range(1, len(headers) + 1):
            value = "\n　　　　　(電話)" if col_idx == 3 else ""
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # 列幅の調整
    column_widths = [15, 10, 50, 20, 20, 20, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=5, column=col_idx).column_letter].width = width

    # メモリ内でファイルを保存
    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exported_file.xlsx"},
    )
